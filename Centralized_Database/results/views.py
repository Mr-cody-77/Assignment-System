from rest_framework.decorators import (
    api_view,
    permission_classes
)

from rest_framework.permissions import AllowAny, IsAuthenticated

from rest_framework.response import Response

from users.permissions import (
    IsTeacher,
    IsStudent
)

from .serializers import (
    PushResultSerializer,
    ResultSerializer
)

from .services import store_result
from .models import Result
from users.models import User

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsTeacher])
def export_results_excel(request):
    """
    Export results as an Excel sheet containing Student Name, Roll Number, Department, Score, etc.
    """
    test_id = request.query_params.get('test_id')
    
    queryset = Result.objects.select_related('student').all()
    
    # Map question_id (string) to test_id to avoid N+1 queries
    from questions.models import Question
    q_map = {str(q.id): q.test_id for q in Question.objects.all()}
    
    # Map plagiarism incidents: (student_id, question_id_str) -> max_similarity_score
    from results.models import PlagiarismDetected, SubmittedSolution
    plag_map = {}
    for p in PlagiarismDetected.objects.all():
        key = (p.flagged_student_id_id, str(p.question_id))
        if key not in plag_map or p.similarity_score > plag_map[key]:
            plag_map[key] = p.similarity_score
            
    # Map plagiarism checked status: (roll_number, question_id_str) -> bool
    submission_checked_map = {}
    for sub in SubmittedSolution.objects.all():
        submission_checked_map[(sub.roll_number, str(sub.question_id))] = sub.plagiarism_checked
            
    # Filter by test_id if provided (in python since question_id is a string field)
    if test_id:
        queryset = [r for r in queryset if str(q_map.get(str(r.question_id))) == str(test_id)]
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Results"
    
    # Define headers
    headers = [
        "Student Name", 
        "Roll Number", 
        "Department", 
        "Test ID", 
        "Question ID", 
        "Score", 
        "Plagiarism Score", 
        "Flagged", 
        "Submitted At"
    ]
    ws.append(headers)
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    plagiarism_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Populate data
    for result in queryset:
        student = result.student
        test_val = q_map.get(str(result.question_id), 'Unknown')
        
        is_checked = submission_checked_map.get((student.roll_number, str(result.question_id)), False)
        
        if not is_checked:
            plag_score_text = "Plagiarism detection is under process"
            flagged_text = "Under process"
            is_flagged = False
        else:
            # Determine plagiarism for this student's question submission
            plag_score = plag_map.get((student.id, str(result.question_id)), 0.0)
            is_flagged = plag_score > 0.0
            
            plag_score_text = f"{plag_score * 100:.2f}%" if is_flagged else "0%"
            flagged_text = "Yes" if is_flagged else "No"
        
        row = [
            student.name or 'N/A',
            student.roll_number,
            student.department or 'N/A',
            test_val,
            result.question_id,
            result.score,
            plag_score_text,
            flagged_text,
            result.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if result.submitted_at else "N/A"
        ]
        ws.append(row)
        
        if is_flagged:
            for cell in ws[ws.max_row]:
                cell.fill = plagiarism_fill
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        
    # Generate response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"results_export_{test_id if test_id else 'all'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response

@api_view(['POST'])
@permission_classes([AllowAny])
def push_result(request):
    serializer = PushResultSerializer(data=request.data)

    if not serializer.is_valid():
        print(serializer.errors)
        return Response(serializer.errors, status=400)

    data = serializer.validated_data

    try:
        student = User.objects.get(roll_number=data["roll_number"])
        if data.get("email"):
            student.email = data["email"]
            student.save()
    except User.DoesNotExist:
        return Response({
            "success": False,
            "message": "Student not found"
        }, status=404)

    result = store_result(student, data)

    return Response({
        "success": True,
        "result_id": result.id
    })

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def get_results(request):

    roll_number = request.data.get('roll_number')

    # Teacher -> can see everything
    if request.user.role == 'teacher':

        if roll_number:
            queryset = Result.objects.filter(roll_number=roll_number)
        else:
            queryset = Result.objects.all()

    # Student -> only own results
    else:

        if roll_number != request.user.roll_number:
            return Response(
                {
                    "success": False,
                    "message": "You can only view your own results"
                },
                status=403
            )

        queryset = Result.objects.filter(
            student=request.user
        )

    serializer = ResultSerializer(
        queryset,
        many=True
    )

    return Response(serializer.data)


# ─────────────────────────────────────────────────────────────────────────────
# PLAGIARISM DETECTION
# ─────────────────────────────────────────────────────────────────────────────

import os
import logging

from .models import SubmittedSolution, SolutionFingerprint, PlagiarismDetected, CodeSubmissionHistory
from .serializers import (
    SolutionIngestSerializer,
    PlagiarismDetectedTeacherSerializer,
    PlagiarismDetectedStudentSerializer,
    CodeSubmissionHistorySerializer,
)
from .plagiarism_engine import build_fingerprint, jaccard_similarity

plagiarism_logger = logging.getLogger("plagiarism.views")


def _get_threshold() -> float:
    """
    Read the similarity threshold from the environment at call-time
    so that changing PLAGIARISM_THRESHOLD in .env takes effect on
    the next request without a code deploy.
    Defaults to 0.75 if the variable is missing or malformed.
    """
    try:
        return float(os.environ.get("PLAGIARISM_THRESHOLD", "0.75"))
    except (TypeError, ValueError):
        plagiarism_logger.warning(
            "PLAGIARISM_THRESHOLD env var is not a valid float; defaulting to 0.75"
        )
        return 0.75


@api_view(['POST'])
@permission_classes([AllowAny])
def plagiarism_ingest(request):
    """
    Internal endpoint called by the Backend Node's async plagiarism pipeline.
    Accepts a student's solution, generates a fingerprint, compares against all
    existing fingerprints for the same question, and logs any matches that
    exceed the configured threshold.

    AllowAny is intentional — this endpoint is called machine-to-machine from
    the Backend Node (which shares no auth context with the Central DB).
    It only writes to the new plagiarism tables, never to existing tables.
    """
    serializer = SolutionIngestSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=400)

    data = serializer.validated_data
    roll_number = data["roll_number"]
    question_id = str(data["question_id"])
    language = data["language"]
    code = data["code"]

    # ── a) Store / update the raw solution ───────────────────────────────────
    SubmittedSolution.objects.update_or_create(
        roll_number=roll_number,
        question_id=question_id,
        defaults={
            "language": language,
            "code": code,
            "plagiarism_checked": False,
        },
    )

    CodeSubmissionHistory.objects.create(
        roll_number=roll_number,
        question_id=question_id,
        language=language,
        code=code,
    )

    # ── b) Generate fingerprint ───────────────────────────────────────────────
    fp_hashes = build_fingerprint(code, language)
    SolutionFingerprint.objects.update_or_create(
        roll_number=roll_number,
        question_id=question_id,
        defaults={
            "language": language,
            "fingerprint_data": fp_hashes,
        },
    )

    if not fp_hashes:
        # Code too short to fingerprint — skip comparison silently
        return Response({"status": "ok", "detail": "fingerprint_too_short"})

    # ── c) Fetch all other fingerprints for the same question ─────────────────
    other_fps = SolutionFingerprint.objects.filter(
        question_id=question_id
    ).exclude(roll_number=roll_number)

    # ── d) Compare & e) Flag if threshold breached ────────────────────────────
    threshold = _get_threshold()
    flagged_count = 0

    for other in other_fps:
        if not other.fingerprint_data:
            continue

        score = jaccard_similarity(fp_hashes, other.fingerprint_data)

        if score >= threshold:
            # Resolve the User object for the currently-submitting student
            try:
                flagged_user = User.objects.get(roll_number=roll_number)
            except User.DoesNotExist:
                plagiarism_logger.warning(
                    "Cannot flag plagiarism: user with roll %s not found.", roll_number
                )
                continue

            PlagiarismDetected.objects.update_or_create(
                flagged_student_id=flagged_user,
                copied_from_student_roll=other.roll_number,
                question_id=question_id,
                defaults={
                    'similarity_score': round(score, 4),
                    'emailed': False
                }
            )
            flagged_count += 1
            plagiarism_logger.info(
                "Plagiarism flagged: %s ~ %s (Q%s, %.1f%%)",
                roll_number, other.roll_number, question_id, score * 100,
            )

    return Response({
        "status": "ok",
        "fingerprint_length": len(fp_hashes),
        "comparisons": other_fps.count(),
        "flagged": flagged_count,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def plagiarism_teacher_view(request):
    """
    Teacher-only: returns all plagiarism flags including both student roll numbers.
    Optionally filter by ?question_id=<id>.
    """
    if request.user.role != 'teacher':
        return Response({"detail": "Permission denied."}, status=403)

    question_id = request.query_params.get("question_id")
    queryset = PlagiarismDetected.objects.select_related("flagged_student_id").all()
    if question_id:
        queryset = queryset.filter(question_id=question_id)

    serializer = PlagiarismDetectedTeacherSerializer(queryset, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def plagiarism_student_view(request):
    """
    Student-only: returns plagiarism flags for the requesting student.
    The copied_from_student_roll field is deliberately excluded from the response.
    """
    if request.user.role != 'student':
        return Response({"detail": "Permission denied."}, status=403)

    queryset = PlagiarismDetected.objects.filter(
        flagged_student_id=request.user
    ).select_related("flagged_student_id")

    serializer = PlagiarismDetectedStudentSerializer(queryset, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsTeacher])
def code_history_view(request):
    """
    Teacher-only: returns the history of code submissions.
    Filter by roll_number and question_id if provided.
    """
    roll_number = request.query_params.get('roll_number')
    question_id = request.query_params.get('question_id')
    
    queryset = CodeSubmissionHistory.objects.all()
    if roll_number:
        queryset = queryset.filter(roll_number=roll_number)
    if question_id:
        queryset = queryset.filter(question_id=question_id)
        
    serializer = CodeSubmissionHistorySerializer(queryset, many=True)
    return Response(serializer.data)
